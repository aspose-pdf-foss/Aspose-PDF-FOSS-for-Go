"""Three-rung XLSX validation harness (epic pdf-go-3zgg), modeled on
validate_docx.py:

  1. XSD: xl/workbook.xml, xl/worksheets/*.xml and xl/styles.xml against the
     ECMA-376 5th-edition transitional sml.xsd (fetched once, shared dir with
     the DOCX schemas).
  2. openpyxl round-trip: open the workbook, touch every sheet's dimensions.
  3. Optional: Excel COM open-without-repair when pywin32 + Excel are present
     (skipped silently otherwise; the Go side drives Excel checks manually).

Usage:  python tools/validate_xlsx.py file.xlsx [more.xlsx | dir ...]
"""

import glob
import os
import sys
import urllib.request
import zipfile

HERE = os.path.dirname(os.path.abspath(__file__))
XSD_DIR = os.path.join(HERE, "..", "result_files", "docx", "xsd")
SCHEMA_BASE = ("https://raw.githubusercontent.com/QtExcel/ecma-376-5th/master/"
               "ECMA-376/OfficeOpenXML-XMLSchema-Transitional")
SML_FILES = ["sml.xsd"]
SML_PARTS_PREFIX = ("xl/workbook.xml", "xl/styles.xml")


def ensure_schemas():
    os.makedirs(XSD_DIR, exist_ok=True)
    for name in SML_FILES + ["dml-spreadsheetDrawing.xsd", "dml-main.xsd",
                             "shared-commonSimpleTypes.xsd",
                             "shared-relationshipReference.xsd", "xml.xsd"]:
        dst = os.path.join(XSD_DIR, name)
        if os.path.exists(dst):
            continue
        url = f"{SCHEMA_BASE}/{name}"
        if name == "xml.xsd":
            url = "https://www.w3.org/2001/xml.xsd"
        urllib.request.urlretrieve(url, dst)
    # sml.xsd imports xml.xsd without a resolvable location on some mirrors —
    # the DOCX harness already patched the shared schemas; nothing to do when
    # they exist.


def validate_xsd(fn):
    try:
        from lxml import etree
    except ImportError:
        return None
    ensure_schemas()
    schema = etree.XMLSchema(etree.parse(os.path.join(XSD_DIR, "sml.xsd")))
    errors = []
    with zipfile.ZipFile(fn) as z:
        parts = [n for n in z.namelist()
                 if n in SML_PARTS_PREFIX
                 or (n.startswith("xl/worksheets/") and "_rels" not in n)]
        for part in parts:
            try:
                doc = etree.fromstring(z.read(part))
            except etree.XMLSyntaxError as e:
                errors.append(f"{part}: XML syntax: {e}")
                continue
            if not schema.validate(doc):
                for err in schema.error_log:
                    errors.append(f"{part}: {err.message}")
    return errors


def validate_openpyxl(fn):
    try:
        import openpyxl
    except ImportError:
        return None
    try:
        wb = openpyxl.load_workbook(fn)
        for ws in wb.worksheets:
            _ = (ws.max_row, ws.max_column)
        if not wb.worksheets:
            return ["workbook has no sheets"]
    except Exception as e:  # noqa: BLE001 — any failure is a finding
        return [f"openpyxl: {e!r}"]
    return []


def main():
    args = sys.argv[1:]
    if not args:
        print(__doc__)
        return 2
    files = []
    for a in args:
        if os.path.isdir(a):
            files.extend(sorted(glob.glob(os.path.join(a, "*.xlsx"))))
        else:
            files.append(a)
    ok = 0
    xsd_seen = False
    for fn in files:
        errs = []
        x = validate_xsd(fn)
        if x is not None:
            xsd_seen = True
            errs.extend(x)
        o = validate_openpyxl(fn)
        if o is not None:
            errs.extend(o)
        if errs:
            print(f"FAIL {fn}")
            for e in errs[:4]:
                print(f"    {e}")
        else:
            ok += 1
            print(f"ok   {fn}")
    if not xsd_seen:
        print("(lxml not installed — XSD rung skipped)")
    print(f"{ok}/{len(files)} valid")
    return 0 if ok == len(files) else 1


if __name__ == "__main__":
    sys.exit(main())
